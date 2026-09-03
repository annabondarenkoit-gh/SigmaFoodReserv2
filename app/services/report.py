"""
Генерація .xlsx-звіту (TC-07) на основі аркуша «Для кухні».

Формує файл із двома листами:
  1. "Зведення для кухні" — day | category | dish | кількість порцій
  2. "Індивідуальний вибір" — ПІБ учня | день | категорія | страва
"""
import io
import json

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app import config


def build_report(sheets, week: int) -> bytes:
    """
    sheets: екземпляр SheetsService
    Повертає байти .xlsx-файлу.
    """
    # актуалізуємо аркуш кухні перед звітом
    counter = sheets.rebuild_kitchen(week)

    wb = Workbook()

    # --- Лист 1: зведення для кухні ---
    ws1 = wb.active
    ws1.title = "Зведення для кухні"
    header_fill = PatternFill("solid", fgColor="2E7D32")
    header_font = Font(bold=True, color="FFFFFF")

    ws1.append(["День", "Категорія", "Страва", "Кількість порцій"])
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # counter: {(day, category, dish): qty}
    def day_order(item):
        day = item[0][0]
        return (config.WEEKDAYS.index(day) if day in config.WEEKDAYS else 99, item[0][1])

    for (day, category, dish), qty in sorted(counter.items(), key=day_order):
        ws1.append([day, category, dish, qty])

    for col, width in zip("ABCD", (14, 12, 45, 16)):
        ws1.column_dimensions[col].width = width

    # --- Лист 2: індивідуальний вибір ---
    ws2 = wb.create_sheet("Індивідуальний вибір")
    ws2.append(["Учень (user_id)", "День", "Категорія", "Страва", "Змінено персоналом"])
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font

    orders = sheets.get_orders_for_week(week)
    for o in orders:
        try:
            dishes = json.loads(o["dishes_json"])
        except (json.JSONDecodeError, TypeError):
            continue
        staff = "так" if str(o.get("is_staff_modified")).lower() == "true" else ""
        for category, items in dishes.items():
            for dish in items:
                ws2.append([o["user_id"], o["day"], category, dish, staff])

    for col, width in zip("ABCDE", (18, 14, 12, 45, 18)):
        ws2.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
